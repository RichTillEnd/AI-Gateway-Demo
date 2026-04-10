"""
RAG (Retrieval-Augmented Generation) 服務
支援上傳公司文件，並在對話時自動檢索相關內容輔助 AI 回答

安裝依賴：
    pip install chromadb openai tiktoken pypdf2 python-docx
"""

import os
import uuid
import json
import hashlib
from typing import Optional
from datetime import datetime

import chromadb
from chromadb.config import Settings
from openai import OpenAI

# ── 設定 ──────────────────────────────────────────────
CHROMA_DB_PATH = "./chroma_db"          # ChromaDB 本地儲存路徑
COLLECTION_NAME = "company_knowledge"   # 知識庫集合名稱
CHUNK_SIZE = 400                        # 每個片段的最大字元數
CHUNK_OVERLAP = 80                      # 片段之間的重疊字元數（避免切斷語意）
TOP_K = 4                               # 最終給 AI 的片段數量
RERANK_TOP_K = 10                       # Re-ranking 前先多撈幾個候選
EMBEDDING_MODEL = "text-embedding-3-small"   # OpenAI Embedding 模型
SIMILARITY_THRESHOLD = 0.55            # 相似度門檻（低於此值不採用）
RERANK_MODEL = "gpt-4.1-mini"           # Re-ranking 使用的模型（輕量快速）


# ── ChromaDB 初始化 ────────────────────────────────────
def get_chroma_client():
    """取得 ChromaDB 客戶端（本地持久化）"""
    client = chromadb.PersistentClient(
        path=CHROMA_DB_PATH,
        settings=Settings(anonymized_telemetry=False)
    )
    return client


def get_collection():
    """取得或建立知識庫集合"""
    client = get_chroma_client()
    collection = client.get_or_create_collection(
        name=COLLECTION_NAME,
        metadata={"hnsw:space": "cosine"}  # 使用 cosine 相似度
    )
    return collection


# ── 文字切割 ───────────────────────────────────────────
def split_text(text: str, chunk_size: int = CHUNK_SIZE, overlap: int = CHUNK_OVERLAP) -> list[str]:
    """
    將長文字切割成小片段
    
    策略：
    1. 先嘗試依段落（換行）切割，保留語意完整性
    2. 若單一段落過長，再依句子切割
    3. 加入重疊，避免切斷語意
    """
    # 先依雙換行（段落）切割
    paragraphs = [p.strip() for p in text.split("\n\n") if p.strip()]
    
    chunks = []
    current_chunk = ""
    
    for para in paragraphs:
        # 如果加入這段後不超過限制，直接合併
        if len(current_chunk) + len(para) + 2 <= chunk_size:
            current_chunk = (current_chunk + "\n\n" + para).strip()
        else:
            # 儲存當前 chunk
            if current_chunk:
                chunks.append(current_chunk)
            
            # 如果單一段落就超過 chunk_size，需要再切割
            if len(para) > chunk_size:
                # 依句子切割（中英文標點）
                sentences = []
                temp = ""
                for char in para:
                    temp += char
                    if char in "。！？.!?\n":
                        sentences.append(temp.strip())
                        temp = ""
                if temp.strip():
                    sentences.append(temp.strip())
                
                sub_chunk = ""
                for sent in sentences:
                    if len(sub_chunk) + len(sent) <= chunk_size:
                        sub_chunk += sent
                    else:
                        if sub_chunk:
                            chunks.append(sub_chunk)
                        sub_chunk = sent
                if sub_chunk:
                    current_chunk = sub_chunk
                else:
                    current_chunk = ""
            else:
                current_chunk = para
    
    if current_chunk:
        chunks.append(current_chunk)
    
    # 加入重疊（將前一個 chunk 的尾部加到當前 chunk 開頭）
    if overlap > 0 and len(chunks) > 1:
        overlapped_chunks = [chunks[0]]
        for i in range(1, len(chunks)):
            prev_tail = chunks[i - 1][-overlap:]
            overlapped_chunks.append(prev_tail + " " + chunks[i])
        return overlapped_chunks
    
    return chunks


# ── Excel 專用：每列獨立成一個片段 ────────────────────────
def extract_excel_rows(file_path: str, filename: str) -> list[str]:
    """
    將 Excel 每一列資料轉成獨立片段清單
    每列格式：「欄位名: 值｜欄位名: 值｜...」
    這樣存入 ChromaDB 時每筆資料是獨立片段，不會被 chunking 切斷
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    all_rows_text = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            continue

        # 自動偵測表頭：找非空欄位數最多的那列
        header_idx = max(range(len(all_rows)),
                         key=lambda i: sum(1 for c in all_rows[i] if c is not None))
        headers = [str(c).strip() if c is not None else "" for c in all_rows[header_idx]]

        for row in all_rows[header_idx + 1:]:
            pairs = []
            for h, v in zip(headers, row):
                if v is not None and str(v).strip():
                    label = h if h else "欄位"
                    pairs.append(f"{label}: {str(v).strip()}")
            if pairs:
                # 加上工作表名稱方便搜尋
                row_text = f"【{sheet_name}】" + "｜".join(pairs)
                all_rows_text.append(row_text)

    wb.close()
    return all_rows_text


# ── 文件解析 ───────────────────────────────────────────
def extract_text_from_file(file_path: str, filename: str) -> str:
    """從各種格式的檔案提取純文字"""
    ext = filename.lower().rsplit(".", 1)[-1]
    
    if ext == "txt":
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    
    elif ext == "pdf":
        # 優先用 pymupdf（較快），退回 PyPDF2
        try:
            import fitz  # pymupdf
            doc = fitz.open(file_path)
            
            # 檢查是否加密
            if doc.is_encrypted:
                # 嘗試空密碼解密
                if not doc.authenticate(""):
                    raise ValueError("此 PDF 有密碼保護，請先移除密碼再上傳")
            
            pages_text = []
            for page in doc:
                t = page.get_text()
                if t.strip():
                    pages_text.append(t)
            doc.close()
            
            text = "\n".join(pages_text)
            if not text.strip():
                raise ValueError("此 PDF 為純圖片掃描檔，無法提取文字（請使用有文字層的 PDF）")
            return text
            
        except ImportError:
            # 退回 PyPDF2
            try:
                import PyPDF2
                text = ""
                with open(file_path, "rb") as f:
                    reader = PyPDF2.PdfReader(f)
                    if reader.is_encrypted:
                        raise ValueError("此 PDF 有密碼保護，請先移除密碼再上傳")
                    for page in reader.pages:
                        t = page.extract_text()
                        if t:
                            text += t + "\n"
                if not text.strip():
                    raise ValueError("此 PDF 為純圖片掃描檔，無法提取文字")
                return text
            except ImportError:
                raise ImportError("請安裝 PDF 解析套件：pip install pymupdf 或 pip install PyPDF2")
    
    elif ext in ("docx", "doc"):
        try:
            import docx
            doc = docx.Document(file_path)
            return "\n\n".join([para.text for para in doc.paragraphs if para.text.strip()])
        except ImportError:
            raise ImportError("請安裝 python-docx：pip install python-docx")
    
    elif ext == "md":
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    
    elif ext in ("xlsx", "xls", "xlsm", "xlsb"):
        try:
            import openpyxl
            import pandas as pd

            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
            sheet_texts = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # 讀出所有有值的列（跳過完全空白的列）
                rows = []
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c).strip() if c is not None else "" for c in row]
                    if any(cells):  # 非全空白列才保留
                        rows.append(cells)

                if not rows:
                    continue

                # 第一列視為表頭
                headers = rows[0]
                data_rows = rows[1:]

                lines = [f"【工作表：{sheet_name}】"]

                # 每列轉成 "欄位名稱: 值" 的自然語言格式，方便 RAG 搜尋
                for row in data_rows:
                    pairs = []
                    for h, v in zip(headers, row):
                        if v:  # 值非空才輸出
                            label = h if h else "欄位"
                            pairs.append(f"{label}: {v}")
                    if pairs:
                        lines.append("｜".join(pairs))

                sheet_texts.append("\n".join(lines))

            wb.close()

            full_text = "\n\n".join(sheet_texts)
            if not full_text.strip():
                raise ValueError("Excel 檔案內容為空，無資料可擷取")
            return full_text

        except ImportError:
            raise ImportError("請安裝 Excel 解析套件：pip install openpyxl")

    else:
        raise ValueError(f"不支援的檔案格式：{ext}（支援：txt, pdf, docx, md, xlsx）")


# ── Embedding ──────────────────────────────────────────
def get_embeddings(texts: list[str], api_key: str) -> list[list[float]]:
    """
    呼叫 OpenAI API 取得文字的向量表示
    - 依 token 數動態分批，避免單批超過 TPM 限制
    - 遇到 429 Rate Limit 自動 retry（最多 5 次，指數退避）
    """
    import time, re

    client = OpenAI(api_key=api_key)

    # 每批最多送 80,000 tokens（text-embedding-3-small 約每字 1.5 token）
    MAX_TOKENS_PER_BATCH = 80_000
    CHARS_PER_TOKEN = 1.5   # 保守估算（中文偏高）

    # ── 依 token 估算動態分批 ──────────────────────────
    batches = []
    current_batch = []
    current_tokens = 0

    for text in texts:
        estimated = int(len(text) / CHARS_PER_TOKEN) + 1
        if current_batch and current_tokens + estimated > MAX_TOKENS_PER_BATCH:
            batches.append(current_batch)
            current_batch = [text]
            current_tokens = estimated
        else:
            current_batch.append(text)
            current_tokens += estimated

    if current_batch:
        batches.append(current_batch)

    # ── 逐批送出，遇 429 自動 retry ────────────────────
    all_embeddings = []

    for batch_idx, batch in enumerate(batches):
        if batch_idx > 0:
            time.sleep(0.5)   # 批次間主動緩速，避免短時間爆衝

        max_retries = 5
        for attempt in range(max_retries):
            try:
                response = client.embeddings.create(
                    model=EMBEDDING_MODEL,
                    input=batch
                )
                all_embeddings.extend([item.embedding for item in response.data])
                break   # 成功，跳出 retry 迴圈

            except Exception as e:
                err_str = str(e)
                is_rate_limit = "429" in err_str or "rate_limit" in err_str.lower()

                if is_rate_limit and attempt < max_retries - 1:
                    # 從錯誤訊息解析建議等待秒數，找不到就用指數退避
                    wait_match = re.search(r'try again in ([0-9.]+)s', err_str)
                    wait_sec = float(wait_match.group(1)) + 0.5 if wait_match else (2 ** attempt) * 2
                    print(f"[Embedding] Rate limit，等待 {wait_sec:.1f}s 後重試 ({attempt+1}/{max_retries})...")
                    time.sleep(wait_sec)
                else:
                    raise   # 非 429 或已超過重試次數，直接拋出

    return all_embeddings

# ── Re-ranking ────────────────────────────────────────
def rerank_contexts(
    query: str,
    contexts: list[dict],
    api_key: str,
    top_k: int = TOP_K
) -> list[dict]:
    """
    用 LLM 對候選片段做 Re-ranking，找出真正最相關的內容
    
    原理：
    - Embedding 搜尋是「語意相近」，可能撈到主題相關但答案不在裡面的片段
    - Re-ranking 讓 LLM 直接判斷「這段文字能回答這個問題嗎？」，更精準
    
    Args:
        query: 使用者的問題
        contexts: Embedding 搜尋的候選片段（通常 8~10 個）
        api_key: OpenAI API key
        top_k: 最終保留幾個片段
    
    Returns:
        重新排序後的片段（最相關的排前面）
    """
    if not contexts or len(contexts) <= top_k:
        return contexts  # 候選數量不夠多，直接回傳不浪費 API
    
    client = OpenAI(api_key=api_key)
    
    # 組成評分 prompt：讓 LLM 給每個片段打分
    candidates_text = ""
    for i, ctx in enumerate(contexts):
        candidates_text += f"""
[片段 {i}]（來源：{ctx["doc_title"]}）
{ctx["text"][:300]}  
"""
    
    prompt = f"""你是一個資料相關性評分員。
    
使用者問題：「{query}」

以下是從知識庫找到的候選片段，請為每個片段評分（0~10分），判斷它對回答這個問題的有用程度。

評分標準：
- 10分：直接包含問題的答案
- 7~9分：高度相關，有助於回答
- 4~6分：部分相關
- 0~3分：不相關或離題

{candidates_text}

請只回傳 JSON 格式，範例：
{{"scores": [8, 3, 9, 1, 7, 2, 6, 4]}}

注意：scores 陣列長度必須等於片段數量 ({len(contexts)})，順序對應片段 0, 1, 2..."""

    try:
        response = client.chat.completions.create(
            model=RERANK_MODEL,
            messages=[{"role": "user", "content": prompt}],
            response_format={"type": "json_object"},
            temperature=0  # 評分要穩定，不需要創意
        )
        
        import json
        result = json.loads(response.choices[0].message.content)
        scores = result.get("scores", [])
        
        if len(scores) != len(contexts):
            print(f"[Re-rank] 分數數量不符，跳過 re-ranking")
            return contexts[:top_k]
        
        # 將分數附加到 contexts，並依分數排序
        for i, ctx in enumerate(contexts):
            ctx["rerank_score"] = scores[i]
        
        ranked = sorted(contexts, key=lambda x: x.get("rerank_score", 0), reverse=True)
        
        # 只保留 top_k，且過濾掉分數太低的（< 4 分）
        final = [c for c in ranked[:top_k] if c.get("rerank_score", 0) >= 4]
        
        if not final:
            # 如果全部分數都很低，還是回傳最高分的，避免完全沒有參考資料
            final = ranked[:1]
        
        print(f"[Re-rank] {len(contexts)} 個候選 → 保留 {len(final)} 個，分數：{[c.get('rerank_score') for c in final]}")
        return final
        
    except Exception as e:
        print(f"[Re-rank] 失敗，退回原始排序：{e}")
        return contexts[:top_k]  # Re-ranking 失敗時退回原本的結果，不影響主流程


# ── 文件管理 ───────────────────────────────────────────
class RAGService:
    """
    RAG 核心服務類別
    負責文件的新增、刪除、查詢
    """
    
    def __init__(self, openai_api_key: str):
        self.api_key = openai_api_key
        self.collection = get_collection()
    
    def add_document(
        self,
        file_path: str,
        filename: str,
        doc_title: str = "",
        uploaded_by: int = None,
        category: str = "general",
        stored_path: str = ""
    ) -> dict:
        """
        新增文件到知識庫
        
        Args:
            file_path: 檔案在伺服器上的路徑
            filename: 原始檔名
            doc_title: 文件標題（顯示用）
            uploaded_by: 上傳者的 user_id
            category: 文件分類（hr/policy/technical/general）
        
        Returns:
            dict: { success, doc_id, chunks_count, message }
        """
        # 產生文件唯一 ID
        doc_id = hashlib.md5(f"{filename}{datetime.utcnow()}".encode()).hexdigest()[:12]
        
        try:
            ext = filename.lower().rsplit(".", 1)[-1]

            # Excel 檔案：每列資料獨立成一個片段，不經過 chunking
            # 避免多筆資料被合併後切割，確保每筆都能被 RAG 搜尋找到
            if ext in ("xlsx", "xls", "xlsm", "xlsb"):
                chunks = extract_excel_rows(file_path, filename)
                if not chunks:
                    return {"success": False, "message": "Excel 檔案無有效資料列"}
            else:
                # 1. 提取文字
                raw_text = extract_text_from_file(file_path, filename)
                if not raw_text.strip():
                    return {"success": False, "message": "文件內容為空"}
                # 2. 切割文字
                chunks = split_text(raw_text)
                if not chunks:
                    return {"success": False, "message": "文件切割失敗"}

            # 3. 取得 Embedding
            embeddings = get_embeddings(chunks, self.api_key)
            
            # 4. 準備 metadata
            metadatas = []
            ids = []
            for i, chunk in enumerate(chunks):
                chunk_id = f"{doc_id}_chunk_{i}"
                ids.append(chunk_id)
                metadatas.append({
                    "doc_id": doc_id,
                    "doc_title": doc_title or filename,
                    "filename": filename,
                    "category": category,
                    "chunk_index": i,
                    "total_chunks": len(chunks),
                    "uploaded_by": str(uploaded_by or ""),
                    "uploaded_at": datetime.utcnow().isoformat(),
                    "stored_path": stored_path or ""
                })
            
            # 5. 分批存入 ChromaDB（單次上限 5461 筆）
            CHROMA_BATCH = 5000
            for _i in range(0, len(ids), CHROMA_BATCH):
                self.collection.add(
                    ids=ids[_i:_i + CHROMA_BATCH],
                    embeddings=embeddings[_i:_i + CHROMA_BATCH],
                    documents=chunks[_i:_i + CHROMA_BATCH],
                    metadatas=metadatas[_i:_i + CHROMA_BATCH]
                )
            
            return {
                "success": True,
                "doc_id": doc_id,
                "chunks_count": len(chunks),
                "message": f"成功新增文件「{doc_title or filename}」，共 {len(chunks)} 個片段"
            }
        
        except Exception as e:
            return {"success": False, "message": f"新增文件失敗：{str(e)}"}
    
    def add_text(
        self,
        text: str,
        doc_title: str,
        category: str = "general",
        uploaded_by: int = None
    ) -> dict:
        """
        直接新增純文字到知識庫（不需要上傳檔案）
        適合直接貼上規定條文、公告等
        """
        doc_id = hashlib.md5(f"{doc_title}{datetime.utcnow()}".encode()).hexdigest()[:12]
        
        try:
            chunks = split_text(text)
            embeddings = get_embeddings(chunks, self.api_key)
            
            ids = [f"{doc_id}_chunk_{i}" for i in range(len(chunks))]
            metadatas = [{
                "doc_id": doc_id,
                "doc_title": doc_title,
                "filename": f"{doc_title}.txt",
                "category": category,
                "chunk_index": i,
                "total_chunks": len(chunks),
                "uploaded_by": str(uploaded_by or ""),
                "uploaded_at": datetime.utcnow().isoformat()
            } for i in range(len(chunks))]
            
            self.collection.add(
                ids=ids,
                embeddings=embeddings,
                documents=chunks,
                metadatas=metadatas
            )
            
            return {
                "success": True,
                "doc_id": doc_id,
                "chunks_count": len(chunks),
                "message": f"成功新增「{doc_title}」，共 {len(chunks)} 個片段"
            }
        except Exception as e:
            return {"success": False, "message": f"新增失敗：{str(e)}"}
    
    def search(self, query: str, top_k: int = TOP_K, category: str = None, use_rerank: bool = True) -> list[dict]:
        """
        查詢與問題最相關的知識片段
        
        流程：
        1. Embedding 搜尋：先撈 RERANK_TOP_K（預設10）個語意相近的候選
        2. Re-ranking：用 LLM 重新評分，篩選出真正能回答問題的片段
        3. 回傳最終 top_k 個片段給 AI
        
        Args:
            query: 使用者的問題
            top_k: 最終返回的片段數量（re-ranking 後）
            category: 限定分類（None 表示搜全部）
            use_rerank: 是否啟用 re-ranking（預設 True）
        
        Returns:
            list of { text, doc_title, filename, category, score, rerank_score }
        """
        query_embedding = get_embeddings([query], self.api_key)[0]
        
        where_filter = {"category": category} if category else None
        
        # Step 1: 先多撈候選（re-ranking 前需要更多選擇）
        candidate_k = RERANK_TOP_K if use_rerank else top_k
        n_results = min(candidate_k, self.collection.count() or 1)
        
        results = self.collection.query(
            query_embeddings=[query_embedding],
            n_results=n_results,
            where=where_filter,
            include=["documents", "metadatas", "distances"]
        )
        
        contexts = []
        if results["documents"] and results["documents"][0]:
            for doc, meta, dist in zip(
                results["documents"][0],
                results["metadatas"][0],
                results["distances"][0]
            ):
                similarity = 1 - dist
                if similarity >= SIMILARITY_THRESHOLD:
                    contexts.append({
                        "text": doc,
                        "doc_title": meta.get("doc_title", "未知文件"),
                        "filename": meta.get("filename", ""),
                        "category": meta.get("category", ""),
                        "score": round(similarity, 4)
                    })
        
        if not contexts:
            return []
        
        # Step 2: Re-ranking（候選夠多才值得做）
        if use_rerank and len(contexts) > top_k:
            contexts = rerank_contexts(
                query=query,
                contexts=contexts,
                api_key=self.api_key,
                top_k=top_k
            )
        else:
            contexts = contexts[:top_k]
        
        return contexts
    
    def delete_document(self, doc_id: str) -> dict:
        """刪除知識庫中的文件（依 doc_id）"""
        try:
            # 取得該文件的所有 chunk ID
            results = self.collection.get(where={"doc_id": doc_id})
            if not results["ids"]:
                return {"success": False, "message": "找不到該文件"}
            
            self.collection.delete(ids=results["ids"])
            return {"success": True, "message": f"已刪除 {len(results['ids'])} 個片段"}
        except Exception as e:
            return {"success": False, "message": str(e)}
    
    def list_documents(self) -> list[dict]:
        """列出知識庫中所有文件（去重複）"""
        try:
            results = self.collection.get(include=["metadatas"])
            seen_docs = {}
            for meta in results["metadatas"]:
                doc_id = meta.get("doc_id")
                if doc_id not in seen_docs:
                    stored_path = meta.get("stored_path", "")
                    seen_docs[doc_id] = {
                        "doc_id": doc_id,
                        "doc_title": meta.get("doc_title"),
                        "filename": meta.get("filename"),
                        "category": meta.get("category"),
                        "uploaded_at": meta.get("uploaded_at"),
                        "total_chunks": meta.get("total_chunks"),
                        "stored_path": stored_path,
                        "has_original": bool(stored_path and os.path.exists(stored_path))
                    }
            return list(seen_docs.values())
        except Exception:
            return []

    def get_document_chunks(self, doc_id: str) -> list[str]:
        """取得文件的所有文字片段（依 chunk_index 排序）"""
        try:
            results = self.collection.get(
                where={"doc_id": doc_id},
                include=["documents", "metadatas"]
            )
            if not results["ids"]:
                return []
            pairs = sorted(
                zip(results["metadatas"], results["documents"]),
                key=lambda x: x[0].get("chunk_index", 0)
            )
            return [doc for _, doc in pairs]
        except Exception:
            return []
    
    def get_stats(self) -> dict:
        """取得知識庫統計資訊"""
        try:
            total_chunks = self.collection.count()
            docs = self.list_documents()
            return {
                "total_documents": len(docs),
                "total_chunks": total_chunks,
                "documents": docs
            }
        except Exception:
            return {"total_documents": 0, "total_chunks": 0, "documents": []}


# ── RAG 輔助 Prompt 組裝 ───────────────────────────────
def build_rag_prompt(user_question: str, contexts: list[dict]) -> str:
    """
    將檢索到的相關文件片段組成 system prompt 的一部分
    
    這段 prompt 會附加在 system prompt 後面，讓 AI 優先參考這些資料
    """
    if not contexts:
        return ""
    
    context_text = "\n\n".join([
        f"【來源：{ctx['doc_title']}（相關度 {ctx['score']:.0%}）】\n{ctx['text']}"
        for ctx in contexts
    ])
    
    return f"""
---
## 參考資料（請優先根據以下內容回答）

{context_text}

---
回答規則：
1. 若問題可以從上方參考資料找到答案，請根據資料內容回答，並說明來源文件名稱
2. 若參考資料不足以回答，請告知使用者並說明你是根據一般知識回答
3. 不要捏造資料中沒有的規定或數字
"""
