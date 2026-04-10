"""
管理員 API 路由
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy.exc import SQLAlchemyError
from typing import Optional, List
from datetime import datetime
from database import (
    get_db, User, Conversation, Message, UsageLog, ErrorLog,
    UserQuota, RateLimit, ApiKey, AuditLog, CustomQA, PromptTemplate, UserMemory,
    RefreshToken, ProjectDocument, Project
)
from auth import hash_password, validate_password_strength, validate_username, validate_email_domain
from password_policy import force_user_password_change
from quota_manager import get_quota_status
from pii_detector import load_config as load_pii_config, save_config as save_pii_config
from core import (
    get_current_user, get_admin_user, log_error, log_audit,
    _hash_api_key, _json,
    CustomQACreate, CustomQAUpdate,
)

router = APIRouter()


@router.get("/admin/users", tags=["管理員"], summary="取得所有用戶列表（僅管理員）")
async def get_all_users(
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """
    獲取所有用戶列表（僅管理員）
    """
    users = db.query(User).all()
    return users


@router.put("/admin/users/{user_id}", tags=["管理員"], summary="更新用戶資訊（僅管理員）")
async def update_user(
    user_id: int,
    user_data: dict,
    request: Request,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """
    更新用戶資訊（僅管理員）

    接收 JSON body:
    {
        "email": "new@email.com",
        "full_name": "新名字",
        "department": "新部門",
        "is_active": true,
        "is_admin": false
    }
    """
    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        raise HTTPException(status_code=404, detail="用戶不存在")

    # 更新欄位
    if "email" in user_data and user_data["email"] is not None:
        # 檢查 email 是否已被使用
        existing = db.query(User).filter(
            User.email == user_data["email"],
            User.id != user_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail="Email 已被使用")
        user.email = user_data["email"]

    if "full_name" in user_data and user_data["full_name"] is not None:
        user.full_name = user_data["full_name"]

    if "department" in user_data and user_data["department"] is not None:
        user.department = user_data["department"]

    if "is_active" in user_data and user_data["is_active"] is not None:
        user.is_active = user_data["is_active"]

    if "is_admin" in user_data and user_data["is_admin"] is not None:
        user.is_admin = user_data["is_admin"]

    db.commit()
    db.refresh(user)

    log_audit(db, "USER_UPDATE", actor=admin_user,
              resource_type="user", resource_id=user_id,
              target_user=user,
              details={k: v for k, v in user_data.items() if k != "password"},
              ip_address=request.client.host if request.client else None,
              user_agent=request.headers.get("user-agent"))

    return user


@router.delete("/admin/users/{user_id}", tags=["管理員"], summary="刪除用戶（僅管理員）")
async def delete_user(
    user_id: int,
    request: Request,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """
    刪除用戶（僅管理員）

    警告：此操作將永久刪除用戶及其所有資料（對話、訊息、使用記錄）
    """
    # 不能刪除自己
    if user_id == admin_user.id:
        raise HTTPException(status_code=400, detail="不能刪除自己的帳號")

    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        raise HTTPException(status_code=404, detail="用戶不存在")

    username = user.username
    user_email = user.email

    # PostgreSQL 強制 FK 約束，需在刪除 user 前手動清除無 cascade 的關聯
    # 1. 先把該用戶的對話 project_id 設 NULL，讓後續刪除 projects 不違反 FK
    db.query(Conversation).filter(Conversation.user_id == user_id).update(
        {"project_id": None}, synchronize_session=False)
    # 2. 刪除無 cascade 關係的子表
    db.query(UserQuota).filter(UserQuota.user_id == user_id).delete(synchronize_session=False)
    db.query(RateLimit).filter(RateLimit.user_id == user_id).delete(synchronize_session=False)
    db.query(ApiKey).filter(ApiKey.user_id == user_id).delete(synchronize_session=False)
    db.query(Project).filter(Project.user_id == user_id).delete(synchronize_session=False)
    # 3. nullable FK 設 NULL（保留審計日誌記錄，但解除 user 參照）
    db.query(RefreshToken).filter(RefreshToken.user_id == user_id).delete(synchronize_session=False)
    db.query(ProjectDocument).filter(ProjectDocument.user_id == user_id).delete(synchronize_session=False)
    db.query(AuditLog).filter(AuditLog.actor_id == user_id).update(
        {"actor_id": None}, synchronize_session=False)
    db.query(AuditLog).filter(AuditLog.target_user_id == user_id).update(
        {"target_user_id": None}, synchronize_session=False)
    db.query(ErrorLog).filter(ErrorLog.user_id == user_id).update(
        {"user_id": None}, synchronize_session=False)
    db.query(CustomQA).filter(CustomQA.created_by == user_id).update(
        {"created_by": None}, synchronize_session=False)
    db.query(PromptTemplate).filter(PromptTemplate.created_by == user_id).update(
        {"created_by": None}, synchronize_session=False)

    log_audit(db, "USER_DELETE", actor=admin_user,
              resource_type="user", resource_id=user_id,
              details={"username": username, "email": user_email},
              ip_address=request.client.host if request.client else None,
              user_agent=request.headers.get("user-agent"))

    # 刪除用戶（cascade 會自動刪除 conversations / usage_logs / memories）
    db.delete(user)
    db.commit()

    return {
        "message": f"用戶 {username} 已刪除",
        "deleted_user_id": user_id
    }


@router.post("/admin/users/{user_id}/force-password-change", tags=["管理員"], summary="強制用戶下次登入更換密碼")
async def force_password_change_admin(
    user_id: int,
    request: Request,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """
    強制用戶下次登入時更換密碼（僅管理員）
    """
    user = db.query(User).filter(User.id == user_id).first()

    if not user:
        raise HTTPException(status_code=404, detail="用戶不存在")

    # 標記為強制更換密碼
    force_user_password_change(user, db)

    log_audit(db, "USER_FORCE_PASSWORD_CHANGE", actor=admin_user,
              resource_type="user", resource_id=user_id,
              target_user=user,
              ip_address=request.client.host if request.client else None,
              user_agent=request.headers.get("user-agent"))

    return {
        "message": f"已標記用戶 {user.username} 需要強制更換密碼",
        "user_id": user_id
    }


@router.post("/admin/users/batch", tags=["管理員"], summary="批次操作用戶（啟用/停用/刪除）")
async def batch_user_action(
    payload: dict,
    request: Request,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """
    批次操作用戶

    接收 JSON body:
    {
        "user_ids": [1, 2, 3],
        "action": "activate" | "deactivate" | "delete"
    }
    """
    user_ids: List[int] = payload.get("user_ids", [])
    action: str = payload.get("action", "")

    if not user_ids:
        raise HTTPException(status_code=400, detail="未選擇任何用戶")
    if action not in ("activate", "deactivate", "delete"):
        raise HTTPException(status_code=400, detail="不支援的操作")

    # 排除管理員自己
    if admin_user.id in user_ids:
        raise HTTPException(status_code=400, detail="不能對自己執行批次操作")

    users = db.query(User).filter(User.id.in_(user_ids)).all()
    if not users:
        raise HTTPException(status_code=404, detail="找不到指定用戶")

    # 排除其他管理員
    admin_targets = [u for u in users if u.is_admin]
    if admin_targets:
        raise HTTPException(
            status_code=400,
            detail=f"不能對管理員執行批次操作：{', '.join(u.username for u in admin_targets)}"
        )

    results = {"success": 0, "failed": 0, "details": []}

    if action in ("activate", "deactivate"):
        new_status = action == "activate"
        for user in users:
            user.is_active = new_status
            results["success"] += 1
            results["details"].append({"id": user.id, "username": user.username, "status": "ok"})
        db.commit()
        log_audit(db, "USER_BATCH_UPDATE", actor=admin_user,
                  resource_type="user",
                  details={"action": action, "user_ids": [u.id for u in users]},
                  ip_address=request.client.host if request.client else None,
                  user_agent=request.headers.get("user-agent"))

    elif action == "delete":
        for user in users:
            try:
                uid = user.id
                uname = user.username
                uemail = user.email
                # 同 single delete 的 FK 清理邏輯
                db.query(Conversation).filter(Conversation.user_id == uid).update(
                    {"project_id": None}, synchronize_session=False)
                db.query(UserQuota).filter(UserQuota.user_id == uid).delete(synchronize_session=False)
                db.query(RateLimit).filter(RateLimit.user_id == uid).delete(synchronize_session=False)
                db.query(ApiKey).filter(ApiKey.user_id == uid).delete(synchronize_session=False)
                db.query(Project).filter(Project.user_id == uid).delete(synchronize_session=False)
                db.query(RefreshToken).filter(RefreshToken.user_id == uid).delete(synchronize_session=False)
                db.query(ProjectDocument).filter(ProjectDocument.user_id == uid).delete(synchronize_session=False)
                db.query(AuditLog).filter(AuditLog.actor_id == uid).update(
                    {"actor_id": None}, synchronize_session=False)
                db.query(AuditLog).filter(AuditLog.target_user_id == uid).update(
                    {"target_user_id": None}, synchronize_session=False)
                db.query(ErrorLog).filter(ErrorLog.user_id == uid).update(
                    {"user_id": None}, synchronize_session=False)
                db.query(CustomQA).filter(CustomQA.created_by == uid).update(
                    {"created_by": None}, synchronize_session=False)
                db.query(PromptTemplate).filter(PromptTemplate.created_by == uid).update(
                    {"created_by": None}, synchronize_session=False)
                db.delete(user)
                db.flush()
                results["success"] += 1
                results["details"].append({"id": uid, "username": uname, "status": "ok"})
            except Exception as e:
                results["failed"] += 1
                results["details"].append({"id": uid, "username": uname, "status": "error", "error": str(e)})
        db.commit()
        log_audit(db, "USER_BATCH_DELETE", actor=admin_user,
                  resource_type="user",
                  details={"user_ids": [d["id"] for d in results["details"] if d["status"] == "ok"]},
                  ip_address=request.client.host if request.client else None,
                  user_agent=request.headers.get("user-agent"))

    action_label = {"activate": "啟用", "deactivate": "停用", "delete": "刪除"}[action]
    return {
        "message": f"批次{action_label}完成：成功 {results['success']} 位，失敗 {results['failed']} 位",
        **results
    }


@router.get("/admin/stats", tags=["管理員"], summary="全站使用統計（總 Token、總費用、用戶數）")
async def get_stats(
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """
    獲取系統統計數據（僅管理員）
    """
    total_users = db.query(User).count()
    total_conversations = db.query(Conversation).count()
    total_messages = db.query(Message).count()

    # 最近 30 天的使用統計
    from datetime import timedelta
    thirty_days_ago = datetime.utcnow() - timedelta(days=30)

    recent_usage = db.query(UsageLog).filter(
        UsageLog.created_at >= thirty_days_ago
    ).all()

    total_cost = sum(log.estimated_cost for log in recent_usage)

    return {
        "total_users": total_users,
        "total_conversations": total_conversations,
        "total_messages": total_messages,
        "total_cost_30days": round(total_cost, 2),
        "api_calls_30days": len(recent_usage)
    }


@router.get("/admin/usage-logs", tags=["管理員"], summary="取得 API 使用記錄（可依用戶、供應商篩選）")
async def get_usage_logs(
    limit: int = 100,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """
    獲取使用記錄（僅管理員）
    """
    logs = db.query(UsageLog).order_by(
        UsageLog.created_at.desc()
    ).limit(limit).all()

    # 關聯用戶資訊
    result = []
    for log in logs:
        user = db.query(User).filter(User.id == log.user_id).first()
        result.append({
            "id": log.id,
            "user_id": log.user_id,
            "username": user.username if user else "已刪除用戶",
            "provider": log.provider,
            "model": log.model,
            "input_tokens": log.input_tokens,
            "output_tokens": log.output_tokens,
            "estimated_cost": log.estimated_cost,
            "created_at": log.created_at
        })

    return result


@router.get("/admin/chart-data", tags=["管理員"], summary="費用趨勢圖表資料（依日期、供應商分組）")
async def get_chart_data(
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """
    獲取圖表數據（僅管理員）
    """
    from datetime import timedelta
    from sqlalchemy import func

    # 最近 7 天的訊息統計
    seven_days_ago = datetime.utcnow() - timedelta(days=7)

    # 按日期分組統計訊息數
    daily_messages = db.query(
        func.date(Message.created_at).label('date'),
        func.count(Message.id).label('count')
    ).filter(
        Message.created_at >= seven_days_ago
    ).group_by(
        func.date(Message.created_at)
    ).all()

    # 轉換為字典
    messages_by_date = {str(row.date): row.count for row in daily_messages}

    # 生成最近 7 天的完整日期列表
    daily_stats = []
    for i in range(6, -1, -1):
        date = (datetime.utcnow() - timedelta(days=i)).date()
        daily_stats.append({
            'date': str(date),
            'count': messages_by_date.get(str(date), 0)
        })

    # 統計 provider 使用分布
    provider_stats = db.query(
        UsageLog.provider,
        func.count(UsageLog.id).label('count')
    ).group_by(
        UsageLog.provider
    ).all()

    provider_distribution = {row.provider: row.count for row in provider_stats}

    return {
        "daily_messages": daily_stats,
        "provider_distribution": provider_distribution
    }


# ── ErrorLog 管理員 API ──────────────────────────────────
@router.get("/admin/error-logs", tags=["管理員"], summary="取得系統錯誤日誌（API錯誤、認證錯誤、系統異常）")
async def get_error_logs(
    limit: int = 200,
    error_type: Optional[str] = None,
    unresolved_only: bool = False,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """取得錯誤日誌列表（管理員）。支援 ?error_type= 與 ?unresolved_only=true 過濾。"""
    query = db.query(ErrorLog)
    if error_type:
        query = query.filter(ErrorLog.error_type == error_type)
    if unresolved_only:
        query = query.filter(ErrorLog.is_resolved == False)
    logs = query.order_by(ErrorLog.created_at.desc()).limit(limit).all()
    return [
        {
            "id": log.id,
            "error_type": log.error_type,
            "error_message": log.error_message,
            "error_detail": log.error_detail,
            "stack_trace": log.stack_trace,
            "user_id": log.user_id,
            "endpoint": log.endpoint,
            "method": log.method,
            "is_resolved": log.is_resolved,
            "resolved_at": log.resolved_at.isoformat() if log.resolved_at else None,
            "created_at": log.created_at.isoformat() if log.created_at else None,
        }
        for log in logs
    ]


@router.put("/admin/error-logs/{error_id}/resolve", tags=["管理員"], summary="標記錯誤日誌為已解決")
async def resolve_error_log(
    error_id: int,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """標記錯誤為已解決（管理員）。"""
    log = db.query(ErrorLog).filter(ErrorLog.id == error_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="錯誤記錄不存在")
    log.is_resolved = True
    log.resolved_at = datetime.utcnow()
    log.resolved_by = admin_user.id
    db.commit()
    return {"success": True}


@router.put("/admin/error-logs/{error_id}/unresolve", tags=["管理員"], summary="取消已解決標記")
async def unresolve_error_log(
    error_id: int,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """還原錯誤為未解決（管理員）。"""
    log = db.query(ErrorLog).filter(ErrorLog.id == error_id).first()
    if not log:
        raise HTTPException(status_code=404, detail="錯誤記錄不存在")
    log.is_resolved = False
    log.resolved_at = None
    log.resolved_by = None
    db.commit()
    return {"success": True}


# ==================== 批次建立用戶 API ====================

@router.get("/admin/users/batch-template", tags=["管理員"], summary="下載批次匯入 Excel 範本")
async def download_batch_template(admin_user: User = Depends(get_admin_user)):
    """下載預填欄位的 Excel 範本，供管理員批次建立用戶使用。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "用戶匯入"

    headers = ["username", "email", "full_name", "department", "password"]
    header_notes = ["帳號（必填，英數字3-20碼）", "Email（必填，@example.com）",
                    "姓名（選填）", "部門（選填）", "初始密碼（選填，空白自動產生）"]

    header_fill = PatternFill("solid", fgColor="6D4FC2")
    header_font = Font(color="FFFFFF", bold=True)

    for col, (h, note) in enumerate(zip(headers, header_notes), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(h), len(note) // 2 + 4)
        ws.cell(row=2, column=col, value=note).font = Font(color="888888", italic=True)

    # 範例資料
    examples = [
        ["johndoe", "johndoe@example.com", "王小明", "業務部", ""],
        ["janedoe", "janedoe@example.com", "李小美", "財務部", "MyPass@123"],
    ]
    for r, row in enumerate(examples, 3):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)

    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=user_import_template.xlsx"}
    )


@router.post("/admin/users/batch-import", tags=["管理員"], summary="批次匯入用戶（Excel/CSV）")
async def batch_import_users(
    file: UploadFile = File(...),
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """
    從 Excel (.xlsx) 或 CSV (.csv) 批次建立用戶。
    - 欄位：username, email, full_name(選), department(選), password(選)
    - 管理員建立的帳號直接啟用，免 Email 驗證
    - 密碼欄位空白時自動產生 12 位隨機密碼
    - 首次登入強制更換密碼
    - 回傳：imported(成功數)、skipped(略過數)、errors(失敗明細)
    """
    import openpyxl
    import csv
    import secrets
    import string
    from io import BytesIO, StringIO

    content = await file.read()
    fname = (file.filename or "").lower()

    def auto_password():
        alphabet = string.ascii_letters + string.digits + "!@#$"
        while True:
            pwd = "".join(secrets.choice(alphabet) for _ in range(12))
            if (any(c.isupper() for c in pwd) and any(c.islower() for c in pwd)
                    and any(c.isdigit() for c in pwd)):
                return pwd

    # 解析檔案 → list of dicts
    rows = []
    try:
        if fname.endswith(".csv"):
            text = content.decode("utf-8-sig")
            reader = csv.DictReader(StringIO(text))
            rows = list(reader)
        elif fname.endswith(".xlsx"):
            wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            headers = None
            for r in ws.iter_rows(values_only=True):
                if headers is None:
                    headers = [str(c).strip().lower() if c else "" for c in r]
                    # 跳過說明列（第2列以 "帳號" / "Email" 開頭的 note 行）
                    continue
                if not any(r):
                    continue
                row_dict = {headers[i]: (str(r[i]).strip() if r[i] is not None else "") for i in range(len(headers))}
                # 跳過明顯的說明列（第一欄包含中文或括號）
                first_val = row_dict.get("username", "")
                if any(c > "\x7f" for c in first_val) or first_val.startswith("("):
                    continue
                rows.append(row_dict)
        else:
            raise HTTPException(status_code=400, detail="僅支援 .xlsx 或 .csv 格式")
    except HTTPException:
        raise
    except UnicodeDecodeError as e:
        raise HTTPException(status_code=400, detail=f"CSV 編碼錯誤，請確認檔案為 UTF-8：{e}")
    except (ValueError, KeyError) as e:
        raise HTTPException(status_code=400, detail=f"檔案格式錯誤：{e}")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"檔案解析失敗：{e}")

    imported, skipped = 0, 0
    errors = []
    generated_passwords = []  # 回傳自動產生的密碼給管理員

    for idx, row in enumerate(rows, start=1):
        username = (row.get("username") or "").strip()
        email = (row.get("email") or "").strip()
        full_name = (row.get("full_name") or row.get("fullname") or "").strip()
        department = (row.get("department") or "").strip()
        password = (row.get("password") or "").strip()

        # 跳過完全空白列
        if not username and not email:
            skipped += 1
            continue

        # 必填驗證
        if not username:
            errors.append({"row": idx, "username": username, "email": email, "reason": "username 為必填"}); skipped += 1; continue
        if not email:
            errors.append({"row": idx, "username": username, "email": email, "reason": "email 為必填"}); skipped += 1; continue

        # 格式驗證
        ok, msg = validate_username(username)
        if not ok:
            errors.append({"row": idx, "username": username, "email": email, "reason": msg}); skipped += 1; continue

        ok, msg = validate_email_domain(email)
        if not ok:
            errors.append({"row": idx, "username": username, "email": email, "reason": msg}); skipped += 1; continue

        # 重複檢查
        if db.query(User).filter(User.username == username).first():
            errors.append({"row": idx, "username": username, "email": email, "reason": "帳號已存在"}); skipped += 1; continue
        if db.query(User).filter(User.email == email).first():
            errors.append({"row": idx, "username": username, "email": email, "reason": "Email 已存在"}); skipped += 1; continue

        # 密碼處理
        auto_gen = False
        if not password:
            password = auto_password()
            auto_gen = True
        else:
            ok, msg = validate_password_strength(password)
            if not ok:
                errors.append({"row": idx, "username": username, "email": email, "reason": f"密碼不符規則：{msg}"}); skipped += 1; continue

        # 建立用戶
        new_user = User(
            username=username,
            email=email,
            hashed_password=hash_password(password),
            full_name=full_name or None,
            department=department or None,
            is_active=True,
            is_admin=False,
            email_verified=True,
            force_password_change=True,
        )
        db.add(new_user)
        try:
            db.commit()
            db.refresh(new_user)
            log_audit(db, "USER_CREATE", actor=admin_user,
                      resource_type="user", resource_id=new_user.id,
                      details={"username": username, "email": email,
                               "department": department, "source": "batch_import"})
            if auto_gen:
                generated_passwords.append({"username": username, "email": email, "password": password})
            imported += 1
        except SQLAlchemyError as e:
            db.rollback()
            errors.append({"row": idx, "username": username, "email": email, "reason": f"DB 寫入失敗：{e}"}); skipped += 1

    return {
        "imported": imported,
        "skipped": skipped,
        "total": imported + skipped,
        "errors": errors,
        "generated_passwords": generated_passwords,
    }


# ==================== 管理員配額管理 API ====================

# 管理員查詢指定用戶配額
@router.get("/admin/users/{user_id}/quota", tags=["管理員"], summary="查看指定用戶的配額設定與使用量")
async def admin_get_user_quota(
    user_id: int,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用戶不存在")
    return get_quota_status(user, db)


@router.put("/admin/users/{user_id}/quota", tags=["管理員"], summary="設定用戶月度 Token 與費用上限")
async def admin_set_user_quota(
    user_id: int,
    data: dict,
    request: Request,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """設定用戶每月 token 和成本上限"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用戶不存在")
    token_limit = int(data.get("monthly_token_limit", 0))
    cost_limit  = float(data.get("monthly_cost_limit", 0))
    if token_limit <= 0 or cost_limit <= 0:
        raise HTTPException(status_code=400, detail="請輸入有效的上限數值")
    quota = db.query(UserQuota).filter(UserQuota.user_id == user_id).first()
    if not quota:
        quota = UserQuota(
            user_id=user_id,
            monthly_token_limit=token_limit,
            monthly_cost_limit=cost_limit,
            current_month_tokens=0,
            current_month_cost=0.0,
            total_tokens=0,
            total_cost=0.0,
            is_quota_exceeded=False,
            quota_warning_sent=False,
            last_reset_date=datetime.utcnow(),
        )
        db.add(quota)
    else:
        quota.monthly_token_limit = token_limit
        quota.monthly_cost_limit  = cost_limit
        if quota.current_month_tokens < token_limit and quota.current_month_cost < cost_limit:
            quota.is_quota_exceeded = False
    db.commit()

    log_audit(db, "QUOTA_UPDATE", actor=admin_user,
              resource_type="quota", resource_id=user_id,
              target_user=user,
              details={"monthly_token_limit": token_limit, "monthly_cost_limit": cost_limit},
              ip_address=request.client.host if request.client else None,
              user_agent=request.headers.get("user-agent"))

    return get_quota_status(user, db)


@router.post("/admin/users/{user_id}/quota/reset", tags=["管理員"], summary="重置用戶當月配額使用量")
async def admin_reset_user_quota_v2(
    user_id: int,
    request: Request,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """清空用戶本月使用量，解除超額狀態"""
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="用戶不存在")
    quota = db.query(UserQuota).filter(UserQuota.user_id == user_id).first()
    if quota:
        quota.current_month_tokens = 0
        quota.current_month_cost   = 0.0
        quota.is_quota_exceeded    = False
        quota.quota_warning_sent   = False
        quota.last_reset_date      = datetime.utcnow()
        db.commit()

    log_audit(db, "QUOTA_RESET", actor=admin_user,
              resource_type="quota", resource_id=user_id,
              target_user=user,
              ip_address=request.client.host if request.client else None,
              user_agent=request.headers.get("user-agent"))

    return {"message": f"已重置 {user.username} 的本月配額", "user_id": user_id}


@router.post("/admin/users/{user_id}/rate-limit/reset", tags=["管理員"], summary="重置用戶速率限制計數器")
async def admin_reset_rate_limit(
    user_id: int,
    request: Request,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """清除用戶速率限制計數器"""
    from database import RateLimit
    rate_limit = db.query(RateLimit).filter(RateLimit.user_id == user_id).first()
    if rate_limit:
        rate_limit.requests_last_minute = 0
        rate_limit.requests_last_hour = 0
        rate_limit.requests_today = 0
        rate_limit.minute_reset_time = datetime.utcnow()
        rate_limit.hour_reset_time = datetime.utcnow()
        rate_limit.day_reset_time = datetime.utcnow()
        db.commit()

    log_audit(db, "RATE_LIMIT_RESET", actor=admin_user,
              resource_type="rate_limit", resource_id=user_id,
              ip_address=request.client.host if request.client else None,
              user_agent=request.headers.get("user-agent"))

    return {"message": f"已重置用戶 {user_id} 的速率限制"}

@router.get("/user/quota-status", tags=["配額"], summary="查看自己本月配額使用狀況（Token 數、費用、剩餘量）")
async def get_user_quota_status(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    return get_quota_status(current_user, db)


@router.get("/user/rate-limit-status", tags=["配額"], summary="查看自己的速率限制使用狀況（每分鐘/小時/天剩餘次數）")
async def get_user_rate_limit_status(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    from rate_limiter import get_rate_limit_status
    return get_rate_limit_status(current_user, db)


# ==================== 審計日誌 API ====================

@router.get("/admin/audit-logs", tags=["管理員"], summary="查詢審計日誌（僅管理員）")
async def get_audit_logs(
    actor_id: Optional[int] = None,
    actor_email: Optional[str] = None,
    target_user_id: Optional[int] = None,
    action: Optional[str] = None,
    resource_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    page: int = 1,
    page_size: int = 50,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """
    查詢審計日誌，支援多條件篩選與分頁。
    - action 範例：USER_LOGIN、USER_UPDATE、USER_DELETE、QUOTA_UPDATE、RAG_UPLOAD 等
    - start_date / end_date 格式：YYYY-MM-DD
    """
    query = db.query(AuditLog)

    if actor_id is not None:
        query = query.filter(AuditLog.actor_id == actor_id)
    if actor_email:
        query = query.filter(AuditLog.actor_email.ilike(f"%{actor_email}%"))
    if target_user_id is not None:
        query = query.filter(AuditLog.target_user_id == target_user_id)
    if action:
        query = query.filter(AuditLog.action == action.upper())
    if resource_type:
        query = query.filter(AuditLog.resource_type == resource_type)
    if start_date:
        try:
            query = query.filter(AuditLog.created_at >= datetime.strptime(start_date, "%Y-%m-%d"))
        except ValueError:
            raise HTTPException(status_code=400, detail="start_date 格式錯誤，請使用 YYYY-MM-DD")
    if end_date:
        try:
            from datetime import timedelta
            query = query.filter(AuditLog.created_at < datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1))
        except ValueError:
            raise HTTPException(status_code=400, detail="end_date 格式錯誤，請使用 YYYY-MM-DD")

    total = query.count()
    logs = query.order_by(AuditLog.created_at.desc()).offset((page - 1) * page_size).limit(page_size).all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [
            {
                "id": log.id,
                "action": log.action,
                "resource_type": log.resource_type,
                "resource_id": log.resource_id,
                "actor_id": log.actor_id,
                "actor_email": log.actor_email,
                "target_user_id": log.target_user_id,
                "target_user_email": log.target_user_email,
                "details": _json.loads(log.details) if log.details else None,
                "ip_address": log.ip_address,
                "created_at": log.created_at.isoformat() if log.created_at else None,
            }
            for log in logs
        ]
    }



# ==================== 成本分攤報表（Chargeback）API ====================

@router.get("/admin/chargeback", tags=["管理員"], summary="部門成本分攤報表")
async def get_chargeback(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """
    按部門彙總 AI 使用費用，含用戶明細。
    start_date / end_date 格式：YYYY-MM-DD（預設本月）
    """
    from datetime import timedelta
    from sqlalchemy import func

    # 日期範圍（預設本月）
    now = datetime.utcnow()
    if start_date:
        try:
            dt_start = datetime.strptime(start_date, "%Y-%m-%d")
        except ValueError:
            raise HTTPException(status_code=400, detail="start_date 格式錯誤，請使用 YYYY-MM-DD")
    else:
        dt_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    if end_date:
        try:
            dt_end = datetime.strptime(end_date, "%Y-%m-%d") + timedelta(days=1)
        except ValueError:
            raise HTTPException(status_code=400, detail="end_date 格式錯誤，請使用 YYYY-MM-DD")
    else:
        dt_end = now + timedelta(days=1)

    # 查詢：UsageLog join User，在指定時間範圍內
    rows = (
        db.query(
            User.department,
            User.id.label("user_id"),
            User.username,
            User.full_name,
            func.count(UsageLog.id).label("request_count"),
            func.sum(UsageLog.input_tokens).label("input_tokens"),
            func.sum(UsageLog.output_tokens).label("output_tokens"),
            func.sum(UsageLog.estimated_cost).label("total_cost"),
            UsageLog.provider,
        )
        .join(UsageLog, UsageLog.user_id == User.id)
        .filter(UsageLog.created_at >= dt_start, UsageLog.created_at < dt_end)
        .group_by(User.department, User.id, User.username, User.full_name, UsageLog.provider)
        .all()
    )

    # 組裝：department → user → provider
    dept_map: dict = {}
    for row in rows:
        dept = row.department or "（未設定部門）"
        if dept not in dept_map:
            dept_map[dept] = {"users": {}, "providers": {}}

        uid = row.user_id
        if uid not in dept_map[dept]["users"]:
            dept_map[dept]["users"][uid] = {
                "user_id": uid,
                "username": row.username,
                "full_name": row.full_name or "",
                "request_count": 0,
                "input_tokens": 0,
                "output_tokens": 0,
                "total_cost": 0.0,
                "providers": {},
            }
        u = dept_map[dept]["users"][uid]
        u["request_count"] += row.request_count or 0
        u["input_tokens"]   += row.input_tokens  or 0
        u["output_tokens"]  += row.output_tokens or 0
        u["total_cost"]     += row.total_cost    or 0.0
        u["providers"][row.provider] = u["providers"].get(row.provider, 0.0) + (row.total_cost or 0.0)

        prov = row.provider
        dept_map[dept]["providers"][prov] = dept_map[dept]["providers"].get(prov, 0.0) + (row.total_cost or 0.0)

    # 計算部門合計 & 全局合計
    grand_total_cost   = 0.0
    grand_total_tokens = 0
    grand_requests     = 0
    departments = []
    for dept, data in dept_map.items():
        users_list = sorted(data["users"].values(), key=lambda x: x["total_cost"], reverse=True)
        dept_cost   = sum(u["total_cost"]   for u in users_list)
        dept_tokens = sum(u["input_tokens"] + u["output_tokens"] for u in users_list)
        dept_reqs   = sum(u["request_count"] for u in users_list)
        departments.append({
            "department":    dept,
            "total_cost":    round(dept_cost, 6),
            "total_tokens":  dept_tokens,
            "request_count": dept_reqs,
            "providers":     {k: round(v, 6) for k, v in data["providers"].items()},
            "users":         [{**u, "total_cost": round(u["total_cost"], 6)} for u in users_list],
        })
        grand_total_cost   += dept_cost
        grand_total_tokens += dept_tokens
        grand_requests     += dept_reqs

    departments.sort(key=lambda x: x["total_cost"], reverse=True)

    # 加上佔比
    for d in departments:
        d["cost_pct"] = round(d["total_cost"] / grand_total_cost * 100, 1) if grand_total_cost else 0.0

    return {
        "period": {
            "start": dt_start.strftime("%Y-%m-%d"),
            "end":   (dt_end - timedelta(days=1)).strftime("%Y-%m-%d"),
        },
        "total_cost":    round(grand_total_cost, 6),
        "total_tokens":  grand_total_tokens,
        "total_requests": grand_requests,
        "departments":   departments,
    }


@router.get("/admin/chargeback/export", tags=["管理員"], summary="匯出成本分攤報表 Excel")
async def export_chargeback(
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """匯出 Excel，含「部門彙總」和「用戶明細」兩個工作表。"""
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 取報表資料（重用上面的邏輯）
    data = await get_chargeback(start_date=start_date, end_date=end_date,
                                admin_user=admin_user, db=db)

    wb = openpyxl.Workbook()

    # ── 通用樣式 ──
    header_fill   = PatternFill("solid", fgColor="4F46E5")
    header_font   = Font(bold=True, color="FFFFFF", size=11)
    subhead_fill  = PatternFill("solid", fgColor="EEF2FF")
    subhead_font  = Font(bold=True, color="1E1B4B", size=10)
    center        = Alignment(horizontal="center", vertical="center")
    thin          = Side(style="thin", color="CBD5E1")
    border        = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(cell, fill=None, font=None):
        cell.fill  = fill  or header_fill
        cell.font  = font  or header_font
        cell.alignment = center
        cell.border    = border

    def style_cell(cell, bold=False, align="left"):
        cell.font      = Font(bold=bold, size=10)
        cell.alignment = Alignment(horizontal=align, vertical="center")
        cell.border    = border

    # ═══════════════════════════════════════
    # Sheet 1：部門彙總
    # ═══════════════════════════════════════
    ws1 = wb.active
    ws1.title = "部門彙總"
    ws1.freeze_panes = "A3"

    period_str = f"{data['period']['start']} ～ {data['period']['end']}"
    ws1.merge_cells("A1:H1")
    title_cell = ws1["A1"]
    title_cell.value = f"AI Gateway 成本分攤報表　{period_str}"
    title_cell.font      = Font(bold=True, size=13, color="1E1B4B")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 28

    headers1 = ["部門", "費用 (USD)", "佔比 (%)", "Token 數", "請求次數",
                "OpenAI 費用", "Gemini 費用", "人數"]
    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=2, column=col, value=h)
        style_header(c)
    ws1.row_dimensions[2].height = 20

    for row_idx, dept in enumerate(data["departments"], 3):
        vals = [
            dept["department"],
            dept["total_cost"],
            dept["cost_pct"],
            dept["total_tokens"],
            dept["request_count"],
            dept["providers"].get("openai", 0.0),
            dept["providers"].get("gemini", 0.0),
            len(dept["users"]),
        ]
        for col, v in enumerate(vals, 1):
            c = ws1.cell(row=row_idx, column=col, value=v)
            style_cell(c, align="right" if col > 1 else "left")
            if col in (2, 6, 7):
                c.number_format = "$#,##0.000000"
            elif col == 3:
                c.number_format = "0.0%"
                c.value = dept["cost_pct"] / 100

    # 合計列
    total_row = len(data["departments"]) + 3
    totals = ["合計", data["total_cost"], 1.0, data["total_tokens"],
              data["total_requests"], "", "", ""]
    for col, v in enumerate(totals, 1):
        c = ws1.cell(row=total_row, column=col, value=v if v != "" else None)
        c.fill   = subhead_fill
        c.font   = subhead_font
        c.alignment = Alignment(horizontal="right" if col > 1 else "left", vertical="center")
        c.border = border
        if col == 2:
            c.number_format = "$#,##0.000000"
        elif col == 3:
            c.number_format = "0.0%"

    col_widths1 = [22, 14, 10, 14, 10, 14, 14, 8]
    for i, w in enumerate(col_widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════
    # Sheet 2：用戶明細
    # ═══════════════════════════════════════
    ws2 = wb.create_sheet("用戶明細")
    ws2.freeze_panes = "A2"

    headers2 = ["部門", "用戶名", "姓名", "費用 (USD)", "Token 數",
                "請求次數", "OpenAI 費用", "Gemini 費用"]
    for col, h in enumerate(headers2, 1):
        c = ws2.cell(row=1, column=col, value=h)
        style_header(c)
    ws2.row_dimensions[1].height = 20

    detail_row = 2
    for dept in data["departments"]:
        for u in dept["users"]:
            vals = [
                dept["department"],
                u["username"],
                u["full_name"],
                u["total_cost"],
                u["input_tokens"] + u["output_tokens"],
                u["request_count"],
                u["providers"].get("openai", 0.0),
                u["providers"].get("gemini", 0.0),
            ]
            for col, v in enumerate(vals, 1):
                c = ws2.cell(row=detail_row, column=col, value=v)
                style_cell(c, align="right" if col > 3 else "left")
                if col in (4, 7, 8):
                    c.number_format = "$#,##0.000000"
            detail_row += 1

    col_widths2 = [20, 16, 16, 14, 12, 10, 14, 14]
    for i, w in enumerate(col_widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # 輸出
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"chargeback_{data['period']['start']}_{data['period']['end']}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )



# ==================== 模型效能儀表板 API ====================

@router.get("/admin/model-stats", tags=["管理員"], summary="模型效能統計（請求數、Token、費用、趨勢）")
async def get_model_stats(
    days: int = 30,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    from sqlalchemy import func
    from datetime import timedelta

    since = datetime.utcnow() - timedelta(days=days)

    # ── 各模型彙總 ────────────────────────────────────────────
    rows = db.query(
        UsageLog.provider,
        UsageLog.model,
        func.count(UsageLog.id).label("requests"),
        func.sum(UsageLog.input_tokens).label("input_tokens"),
        func.sum(UsageLog.output_tokens).label("output_tokens"),
        func.sum(UsageLog.estimated_cost).label("total_cost"),
    ).filter(
        UsageLog.created_at >= since
    ).group_by(
        UsageLog.provider, UsageLog.model
    ).order_by(func.count(UsageLog.id).desc()).all()

    by_model = [
        {
            "provider":     r.provider,
            "model":        r.model or "(未知)",
            "requests":     r.requests or 0,
            "input_tokens": int(r.input_tokens or 0),
            "output_tokens":int(r.output_tokens or 0),
            "total_tokens": int((r.input_tokens or 0) + (r.output_tokens or 0)),
            "total_cost":   round(float(r.total_cost or 0), 6),
            "avg_tokens":   round(((r.input_tokens or 0) + (r.output_tokens or 0)) / max(r.requests, 1)),
        }
        for r in rows
    ]

    # ── 每日趨勢 ──────────────────────────────────────────────
    daily_rows = db.query(
        func.date(UsageLog.created_at).label("date"),
        func.count(UsageLog.id).label("requests"),
        func.sum(UsageLog.estimated_cost).label("cost"),
        func.sum(UsageLog.input_tokens + UsageLog.output_tokens).label("tokens"),
    ).filter(
        UsageLog.created_at >= since
    ).group_by(
        func.date(UsageLog.created_at)
    ).order_by(func.date(UsageLog.created_at).asc()).all()

    daily_map = {str(r.date): r for r in daily_rows}
    daily_trend = []
    for i in range(days - 1, -1, -1):
        d = str((datetime.utcnow() - timedelta(days=i)).date())
        r = daily_map.get(d)
        daily_trend.append({
            "date":     d,
            "requests": r.requests if r else 0,
            "cost":     round(float(r.cost or 0), 6) if r else 0,
            "tokens":   int(r.tokens or 0) if r else 0,
        })

    # ── 全期總計 ──────────────────────────────────────────────
    totals = db.query(
        func.count(UsageLog.id).label("requests"),
        func.sum(UsageLog.input_tokens).label("input_tokens"),
        func.sum(UsageLog.output_tokens).label("output_tokens"),
        func.sum(UsageLog.estimated_cost).label("total_cost"),
    ).filter(UsageLog.created_at >= since).first()

    return {
        "days":        days,
        "by_model":    by_model,
        "daily_trend": daily_trend,
        "totals": {
            "requests":      totals.requests or 0,
            "input_tokens":  int(totals.input_tokens or 0),
            "output_tokens": int(totals.output_tokens or 0),
            "total_tokens":  int((totals.input_tokens or 0) + (totals.output_tokens or 0)),
            "total_cost":    round(float(totals.total_cost or 0), 6),
        }
    }


# ==================== PII 偵測設定 API ====================

@router.get("/admin/pii-settings", tags=["管理員"], summary="取得 PII 偵測設定")
async def get_pii_settings(admin_user: User = Depends(get_admin_user)):
    return load_pii_config()


@router.put("/admin/pii-settings", tags=["管理員"], summary="更新 PII 偵測設定")
async def update_pii_settings(
    settings: dict,
    admin_user: User = Depends(get_admin_user)
):
    save_pii_config(settings)
    return settings


# ==================== Custom QA 管理 API ====================

@router.get("/admin/custom-qa", tags=["自訂QA"], summary="列出所有自訂 QA 規則（僅管理員）")
async def list_custom_qa(
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """列出所有自訂 QA 規則"""
    rules = db.query(CustomQA).order_by(CustomQA.created_at.desc()).all()
    return [
        {
            "id": r.id,
            "name": r.name,
            "keywords": r.keywords,
            "match_type": r.match_type,
            "answer": r.answer,
            "is_enabled": r.is_enabled,
            "hit_count": r.hit_count or 0,
            "created_at": r.created_at.isoformat() if r.created_at else None,
            "updated_at": r.updated_at.isoformat() if r.updated_at else None,
        }
        for r in rules
    ]


@router.post(
    "/admin/custom-qa",
    tags=["自訂QA"],
    summary="新增自訂 QA 規則",
    description="設定關鍵字命中時直接回傳預設答案，不呼叫 AI API，有效節省費用。支援 any（任一命中）/ all（全部命中）兩種比對模式。",
)
async def create_custom_qa(
    data: CustomQACreate,
    request: Request,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """新增自訂 QA 規則"""
    if data.match_type not in ("any", "all"):
        raise HTTPException(status_code=400, detail="match_type 只接受 'any' 或 'all'")
    rule = CustomQA(
        name=data.name,
        keywords=data.keywords,
        match_type=data.match_type,
        answer=data.answer,
        is_enabled=data.is_enabled,
        created_by=admin_user.id,
    )
    db.add(rule)
    db.commit()
    db.refresh(rule)

    log_audit(db, "QA_CREATE", actor=admin_user,
              resource_type="custom_qa", resource_id=rule.id,
              details={"name": data.name, "keywords": data.keywords, "match_type": data.match_type},
              ip_address=request.client.host if request.client else None,
              user_agent=request.headers.get("user-agent"))

    return {"success": True, "id": rule.id, "message": f"規則「{rule.name}」建立成功"}


@router.put("/admin/custom-qa/{rule_id}", tags=["自訂QA"], summary="更新自訂 QA 規則內容")
async def update_custom_qa(
    rule_id: int,
    data: CustomQAUpdate,
    request: Request,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """更新自訂 QA 規則"""
    rule = db.query(CustomQA).filter(CustomQA.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="規則不存在")
    if data.name is not None:
        rule.name = data.name
    if data.keywords is not None:
        rule.keywords = data.keywords
    if data.match_type is not None:
        if data.match_type not in ("any", "all"):
            raise HTTPException(status_code=400, detail="match_type 只接受 'any' 或 'all'")
        rule.match_type = data.match_type
    if data.answer is not None:
        rule.answer = data.answer
    if data.is_enabled is not None:
        rule.is_enabled = data.is_enabled
    db.commit()

    log_audit(db, "QA_UPDATE", actor=admin_user,
              resource_type="custom_qa", resource_id=rule_id,
              details={k: v for k, v in data.model_dump().items() if v is not None},
              ip_address=request.client.host if request.client else None,
              user_agent=request.headers.get("user-agent"))

    return {"success": True, "message": f"規則「{rule.name}」更新成功"}


@router.delete("/admin/custom-qa/{rule_id}", tags=["自訂QA"], summary="刪除自訂 QA 規則")
async def delete_custom_qa(
    rule_id: int,
    request: Request,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """刪除自訂 QA 規則"""
    rule = db.query(CustomQA).filter(CustomQA.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="規則不存在")
    name = rule.name

    log_audit(db, "QA_DELETE", actor=admin_user,
              resource_type="custom_qa", resource_id=rule_id,
              details={"name": name},
              ip_address=request.client.host if request.client else None,
              user_agent=request.headers.get("user-agent"))

    db.delete(rule)
    db.commit()
    return {"success": True, "message": f"規則「{name}」已刪除"}


@router.patch("/admin/custom-qa/{rule_id}/toggle", tags=["自訂QA"], summary="啟用 / 停用自訂 QA 規則")
async def toggle_custom_qa(
    rule_id: int,
    request: Request,
    admin_user: User = Depends(get_admin_user),
    db: Session = Depends(get_db)
):
    """快速切換規則啟用/停用"""
    rule = db.query(CustomQA).filter(CustomQA.id == rule_id).first()
    if not rule:
        raise HTTPException(status_code=404, detail="規則不存在")
    rule.is_enabled = not rule.is_enabled
    db.commit()

    log_audit(db, "QA_TOGGLE", actor=admin_user,
              resource_type="custom_qa", resource_id=rule_id,
              details={"name": rule.name, "is_enabled": rule.is_enabled},
              ip_address=request.client.host if request.client else None,
              user_agent=request.headers.get("user-agent"))

    status_text = "啟用" if rule.is_enabled else "停用"
    return {"success": True, "is_enabled": rule.is_enabled, "message": f"規則「{rule.name}」已{status_text}"}

# ==================== 旅遊團體 Excel 上傳 ====================

# Excel 欄位名稱 → 資料庫欄位名稱對照表
_EXCEL_COL_MAP = {
    "大分類": "category_major",
    "小分類": "category_minor",
    "出發日": "departure_date",
    "團體代碼": "group_code",
    "團名": "group_name",
    "產品名稱": "product_name",
    "同業價": "trade_price",
    "直客價": "direct_price",
    "總機位": "total_seats",
    "保留": "reserved_seats",
    "報名": "enrolled",
    "可售": "available_seats",
    "收訂": "confirmed_bookings",
    "略訂": "tentative_bookings",
    "訂金": "deposit",
    "班機": "flight",
    "進出點": "entry_exit_point",
    "狀態": "status",
    "團型": "group_type",
    "員工備註": "staff_notes",
}


@router.post("/admin/tour-groups/upload", tags=["旅遊團體"], summary="上傳旅遊團體 Excel（僅管理員）")
async def upload_tour_groups_excel(
    file: UploadFile = File(...),
    admin_user: User = Depends(get_admin_user),
):
    """
    上傳 Excel 檔案，自動匯入 tour_groups 資料表。
    - 僅接受 .xlsx / .xls 格式
    - 每次上傳會完整取代舊資料
    - 欄位對照：大分類→category_major … 員工備註→staff_notes
    """
    import pandas as pd
    import sqlite3 as _sqlite3
    import io
    import os

    # 驗證副檔名
    filename = file.filename or ""
    if not filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="只接受 .xlsx 或 .xls 格式")

    content = await file.read()
    if len(content) > 20 * 1024 * 1024:  # 20 MB 上限
        raise HTTPException(status_code=400, detail="檔案大小不可超過 20 MB")

    try:
        df = pd.read_excel(io.BytesIO(content), dtype=str)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"Excel 格式錯誤：{e}")
    except Exception as e:
        # openpyxl / xlrd 的例外型別因版本不同，保留寬泛 catch
        raise HTTPException(status_code=400, detail=f"Excel 讀取失敗：{e}")

    # 去除欄位名稱前後空白
    df.columns = [str(c).strip() for c in df.columns]

    # 檢查必要欄位
    missing = [c for c in _EXCEL_COL_MAP if c not in df.columns]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Excel 缺少必要欄位：{', '.join(missing)}"
        )

    # 只保留已定義的欄位，並重新命名為英文
    df = df[list(_EXCEL_COL_MAP.keys())].rename(columns=_EXCEL_COL_MAP)

    # 空字串統一轉 None
    df = df.where(df.notna() & (df != ""), other=None)

    # 寫入 SQLite（完整取代）
    db_path = os.getenv("DB_PATH", "./ai_gateway.db")
    conn = _sqlite3.connect(db_path)
    try:
        df.to_sql("tour_groups", conn, if_exists="replace", index=False)
        conn.commit()
    finally:
        conn.close()

    return {
        "success": True,
        "rows_imported": len(df),
        "columns": list(df.columns),
        "message": f"成功匯入 {len(df):,} 筆旅遊團體資料",
    }


@router.get("/admin/tour-groups/info", tags=["旅遊團體"], summary="查詢目前旅遊團體資料狀態（僅管理員）")
async def get_tour_groups_info(admin_user: User = Depends(get_admin_user)):
    """回傳目前 tour_groups 資料表的筆數與最近一筆出發日"""
    import sqlite3 as _sqlite3
    import os
    db_path = os.getenv("DB_PATH", "./ai_gateway.db")
    conn = _sqlite3.connect(db_path)
    try:
        cur = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='tour_groups'")
        if not cur.fetchone():
            return {"exists": False, "row_count": 0, "message": "尚未上傳任何資料"}
        row_count = conn.execute("SELECT COUNT(*) FROM tour_groups").fetchone()[0]
        latest = conn.execute(
            "SELECT MAX(departure_date) FROM tour_groups"
        ).fetchone()[0]
        return {"exists": True, "row_count": row_count, "latest_departure": latest}
    finally:
        conn.close()


# ==================== API Key 管理 ====================

from pydantic import BaseModel
from typing import Optional as Opt

class ApiKeyCreate(BaseModel):
    name: str  # 描述名稱，如 "RPA 機器人"
    expires_days: Opt[int] = None  # 有效天數，None = 永不過期

@router.post("/admin/api-keys", tags=["API Key 管理"], summary="建立新 API Key（僅管理員）")
async def create_api_key(
    payload: ApiKeyCreate,
    request: Request,
    db: Session = Depends(get_db),
    admin_user: User = Depends(get_admin_user)
):
    """產生一組新的 API Key，Key 只在此次回應中顯示一次，請立即複製保存。"""
    import secrets
    from datetime import timedelta

    # 產生 Key：格式 sk-gw-<32位隨機hex>
    raw_key = "sk-gw-" + secrets.token_hex(16)
    key_prefix = raw_key[:12]  # 顯示用前綴，如 sk-gw-ab12cd

    # HMAC-SHA256 with server-side pepper（設定 API_KEY_PEPPER 環境變數可啟用）
    hashed = _hash_api_key(raw_key)

    expires_at = None
    if payload.expires_days:
        expires_at = datetime.utcnow() + timedelta(days=payload.expires_days)

    api_key = ApiKey(
        user_id=admin_user.id,
        name=payload.name,
        key_prefix=key_prefix,
        hashed_key=hashed,
        expires_at=expires_at,
    )
    db.add(api_key)
    db.commit()
    db.refresh(api_key)

    log_audit(db, "API_KEY_CREATE", actor=admin_user,
              resource_type="api_key", resource_id=api_key.id,
              details={"name": payload.name, "key_prefix": key_prefix, "expires_days": payload.expires_days},
              ip_address=request.client.host if request.client else None,
              user_agent=request.headers.get("user-agent"))

    return {
        "id": api_key.id,
        "name": api_key.name,
        "key": raw_key,          # 只回傳一次！
        "key_prefix": key_prefix,
        "expires_at": api_key.expires_at,
        "created_at": api_key.created_at,
        "message": "請立即複製此 Key，關閉後將無法再次查看。"
    }


@router.get("/admin/api-keys", tags=["API Key 管理"], summary="列出所有 API Key（僅管理員）")
async def list_api_keys(
    db: Session = Depends(get_db),
    admin_user: User = Depends(get_admin_user)
):
    """列出所有 API Key（不含完整 Key，只顯示前綴）"""
    keys = db.query(ApiKey).order_by(ApiKey.created_at.desc()).all()
    return [
        {
            "id": k.id,
            "name": k.name,
            "key_prefix": k.key_prefix,
            "is_active": k.is_active,
            "last_used_at": k.last_used_at,
            "expires_at": k.expires_at,
            "created_at": k.created_at,
        }
        for k in keys
    ]


@router.patch("/admin/api-keys/{key_id}/toggle", tags=["API Key 管理"], summary="啟用 / 停用 API Key（僅管理員）")
async def toggle_api_key(
    key_id: int,
    request: Request,
    db: Session = Depends(get_db),
    admin_user: User = Depends(get_admin_user)
):
    key = db.query(ApiKey).filter(ApiKey.id == key_id).first()
    if not key:
        raise HTTPException(status_code=404, detail="API Key 不存在")
    key.is_active = not key.is_active
    db.commit()

    log_audit(db, "API_KEY_TOGGLE", actor=admin_user,
              resource_type="api_key", resource_id=key_id,
              details={"name": key.name, "is_active": key.is_active},
              ip_address=request.client.host if request.client else None,
              user_agent=request.headers.get("user-agent"))

    return {"id": key.id, "name": key.name, "is_active": key.is_active}


@router.delete("/admin/api-keys/{key_id}", tags=["API Key 管理"], summary="刪除 API Key（僅管理員）")
async def delete_api_key(
    key_id: int,
    request: Request,
    db: Session = Depends(get_db),
    admin_user: User = Depends(get_admin_user)
):
    key = db.query(ApiKey).filter(ApiKey.id == key_id).first()
    if not key:
        raise HTTPException(status_code=404, detail="API Key 不存在")
    key_name = key.name
    key_prefix = key.key_prefix

    log_audit(db, "API_KEY_DELETE", actor=admin_user,
              resource_type="api_key", resource_id=key_id,
              details={"name": key_name, "key_prefix": key_prefix},
              ip_address=request.client.host if request.client else None,
              user_agent=request.headers.get("user-agent"))

    db.delete(key)
    db.commit()
    return {"message": f"API Key '{key_name}' 已刪除"}


# ==================== 語意快取管理（管理員）====================

@router.get("/admin/semantic-cache/stats", tags=["管理員"], summary="語意快取統計")
async def get_cache_stats(current_user: User = Depends(get_admin_user)):
    """取得 Redis 語意快取的統計資訊。"""
    from semantic_cache import cache_stats
    return await cache_stats()


@router.delete("/admin/semantic-cache", tags=["管理員"], summary="清除語意快取")
async def clear_semantic_cache(current_user: User = Depends(get_admin_user), db: Session = Depends(get_db)):
    """清除所有快取條目（例如知識庫更新後需要清除）。"""
    from semantic_cache import cache_clear
    deleted = await cache_clear()
    log_audit(db, "CACHE_CLEAR", current_user, resource_type="semantic_cache",
              details={"deleted_entries": deleted})
    return {"deleted": deleted, "message": f"已清除 {deleted} 筆快取"}
